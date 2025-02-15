from django.shortcuts import render
from django.http import HttpResponse
from .forms import ExcelUploadForm
from .models import Student, Grade
import openpyxl
from io import BytesIO


def upload_grades(request):
    if request.method == "POST":
        # Check if the file is present in the request
        excel_file = request.FILES.get("excel_file")

        if excel_file:
            form = ExcelUploadForm(request.POST, request.FILES)
            if form.is_valid():
                workbook = openpyxl.load_workbook(excel_file)
                sheet = workbook.active

                uploaded_data = []
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    student_name, email, subject, grade = row
                    student, created = Student.objects.get_or_create(
                        name=student_name,
                        email=email,
                    )

                    grade_student = Grade.objects.filter(
                        student=student, subject=subject
                    ).first()
                    if grade_student:
                        grade_student.subject = subject
                        grade_student.grade = grade
                        grade_student.save()
                    else:
                        Grade.objects.create(
                            student=student, subject=subject, grade=grade
                        )

                    uploaded_data.append(
                        {
                            "student_name": student_name,
                            "email": email,
                            "subject": subject,
                            "grade": grade,
                        }
                    )

                return render(
                    request, "grades/uploaded_success.html", {"data": uploaded_data}
                )
        else:
            form = ExcelUploadForm()
            return render(
                request,
                "grades/upload_grades.html",
                {"form": form, "error": "No file uploaded."},
            )

    else:
        form = ExcelUploadForm()

    return render(request, "grades/upload_grades.html", {"form": form})


def display_grades(request):
    students = Student.objects.all()
    student_averages = {}

    for student in students:
        grades = Grade.objects.filter(student=student)
        total_grades = sum(grade.grade for grade in grades)
        average_grade = total_grades / len(grades) if grades else 0
        student_averages[student] = average_grade

    class_average = (
        sum(student_averages.values()) / len(student_averages)
        if student_averages
        else 0
    )
    print(student_averages.items())

    return render(
        request,
        "grades/display_grades.html",
        {
            "student_averages": student_averages,
            "class_average": round(class_average, 2),
        },
    )


def export_grades(request):
    students = Student.objects.all()
    grades_data = []

    for student in students:
        grades = Grade.objects.filter(student=student)
        total_grades = sum(grade.grade for grade in grades)
        average_grade = total_grades / len(grades) if grades else 0
        grades_data.append([student.name, student.email, average_grade])

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Student Name", "Email", "Average Grade"])

    for row in grades_data:
        ws.append(row)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="grades_report.xlsx"'
    return response
